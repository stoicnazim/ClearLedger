import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

BG = "0D1117"; SURFACE = "161B22"; INPUT_BG = "1A2332"
HEADER_BLUE = "58A6FF"; EDIT_BLUE = "4488FF"; LABEL = "E6EDF3"
DIM = "8B949E"; NOTE = "484F58"; WHITE = "FFFFFF"
GREEN = "3FB950"; RED = "F85149"; ORANGE = "D29922"

font_body = Font(name="Aptos", size=11, color=LABEL)
font_header14 = Font(name="Aptos", size=14, bold=True, color=HEADER_BLUE)
font_section = Font(name="Aptos", size=12, bold=True, color=HEADER_BLUE)
font_editable = Font(name="Aptos", size=11, color=EDIT_BLUE)
font_dim = Font(name="Aptos", size=10, color=DIM)
font_note = Font(name="Aptos", size=10, color=NOTE)
font_wb = Font(name="Aptos", size=11, bold=True, color=WHITE)
font_white = Font(name="Aptos", size=11, color=WHITE)

fill_bg = PatternFill(start_color=BG, end_color=BG, fill_type="solid")
fill_surface = PatternFill(start_color=SURFACE, end_color=SURFACE, fill_type="solid")
fill_input = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
fill_header = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
fill_red = PatternFill(start_color=RED, end_color=RED, fill_type="solid")

thin_border = Border(left=Side(style="thin",color="30363D"),right=Side(style="thin",color="30363D"),top=Side(style="thin",color="30363D"),bottom=Side(style="thin",color="30363D"))
align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)

TEMPLATES_DIR = r"C:\Users\nazim\Desktop\ClearLedger\templates"

def sc(ws, r, c, v, font=font_body, fill=fill_bg, align=align_l, border=thin_border, nf=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font; cell.fill = fill; cell.alignment = align; cell.border = border
    if nf: cell.number_format = nf
    return cell

def editable(ws, r, c, v="", nf=None):
    return sc(ws, r, c, v, font_editable, fill_input, align_c, nf=nf)

def label(ws, r, c, v):
    return sc(ws, r, c, v, font_body, fill_surface)

def section_hdr(ws, r, c, v, n=1):
    for i in range(c, c+n):
        cell = ws.cell(row=r, column=i)
        cell.font = font_section; cell.fill = fill_surface; cell.alignment = align_l; cell.border = thin_border
    ws.cell(row=r, column=c, value=v)

def header_row(ws, r, vals):
    for i, v in enumerate(vals, 1):
        sc(ws, r, i, v, font_wb, fill_header, align_c)

def col_widths(ws, w):
    for i, ww in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = ww

def new_sheet(wb, title, hidden=False):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = HEADER_BLUE
    if hidden: ws.sheet_state = "hidden"
    col_widths(ws, [3])
    return ws

def save_wb(wb, name):
    fp = f"{TEMPLATES_DIR}\\{name}"
    wb.save(fp)
    return fp

# =====================================================================
# 1. DISPUTE RESOLUTION
# =====================================================================
def build_dispute():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws, [3,35,18,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"Dispute Resolution Process Pack - APQC PCF v8.0",font_header14)
    sc(ws,3,2,"Reference: PCF v8.0 Process 10903 - Manage Disputes",font_dim)
    header_row(ws,5,["","KPI","Current","Target","Benchmark P50","Benchmark P75","Status"])
    kpis = [("Dispute rate (% of invoices)","3.5","1.5","2.0","1.0","Behind"),("Avg resolution time (days)","12","5","7","4","Behind"),("Dispute as % of AR","2.8","1.0","1.5","0.8","Behind"),("First-contact resolution (%)","45","70","60","80","Behind"),("Customer satisfaction score","3.2","4.5","4.0","4.8","Near"),("Cost per dispute (EUR)","85","35","50","30","Near")]
    for i,k in enumerate(kpis,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,k[0])
        for j,v in enumerate(k[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"Dispute Register")
    col_widths(ws2,[3,18,22,18,18,18,18,18,18,18])
    header_row(ws2,1,["","Case #","Customer","Invoice Ref","Amount (EUR)","Reason Code","Date Opened","Age (Days)","Status","Owner"])
    for r in range(2,16):
        sc(ws2,r,1,"",fill=fill_bg)
        for c in range(2,11): editable(ws2,r,c)
    section_hdr(ws2,17,2,"Summary",8)
    rs=18
    for l,r2 in [("Total open disputes","=COUNTA(H2:H16)"),("Total disputed amount","=SUM(E2:E16)"),("Avg resolution age","=AVERAGE(H2:H16)")]:
        label(ws2,rs,2,l); sc(ws2,rs,3,r2,font_editable,fill_input,align_c)
        rs+=1

    ws3 = new_sheet(wb,"Resolution Workflow")
    col_widths(ws3,[3,8,35,20,20,18])
    header_row(ws3,1,["","Step","Action","Owner","SLA (days)","Notes"])
    steps = [("1","Receive & log dispute","AR Team","1",""),("2","Validate & assign reason code","AR Lead","1",""),("3","Gather supporting docs","AR Analyst","2",""),("4","Investigate root cause","Dispute Team","3",""),("5","Propose resolution","Dispute Lead","2",""),("6","Customer approval","Sales/AM","5",""),("7","Apply credit / adjustment","AR Team","1",""),("8","Close & document","AR Analyst","1",""),("9","Root cause analysis","Quality","5","Post-close")]
    for i,(s,a,o,sla,n) in enumerate(steps,2):
        sc(ws3,i,1,"",fill=fill_bg); label(ws3,i,2,s); label(ws3,i,3,a)
        editable(ws3,i,4,o); editable(ws3,i,5,sla); editable(ws3,i,6,n)

    ws4 = new_sheet(wb,"Root Cause Analysis")
    col_widths(ws4,[3,30,18,18,30])
    header_row(ws4,1,["","Root Cause Category","Count","% of Total","Action Plan"])
    causes = [("Pricing / billing error","","",""),("Service delivery issue","","",""),("Contract terms dispute","","",""),("Late delivery / short shipment","","",""),("Quality / specification","","",""),("Invoice error (PO mismatch)","","",""),("Duplicate invoice","","",""),("Other","","","")]
    for i,(cat,_,_,_) in enumerate(causes,2):
        sc(ws4,i,1,"",fill=fill_bg); label(ws4,i,2,cat)
        editable(ws4,i,3); editable(ws4,i,4); editable(ws4,i,5)
    label(ws4,11,2,"Total"); sc(ws4,11,3,"=SUM(C2:C9)",font_editable,fill_input,align_c)
    sc(ws4,11,4,"=SUM(D2:D9)",font_editable,fill_input,align_c)

    ws5 = new_sheet(wb,"Action Plan")
    col_widths(ws5,[3,8,35,20,20,18])
    header_row(ws5,1,["","Phase","Action","Owner","Target Date","Status"])
    items = [("1","Map current dispute workflow","AR Lead",""),("1","Implement reason code taxonomy","Quality",""),("2","Set up automated dispute logging","IT/AR",""),("2","Define SLAs per category","AR Lead",""),("3","Deploy dispute dashboard","IT",""),("3","Monthly RCA review cadence","Quality",""),("4","Optimize and benchmark","AR Lead","")]
    for i,(p,a,o,s) in enumerate(items,2):
        sc(ws5,i,1,"",fill=fill_bg); label(ws5,i,2,p); label(ws5,i,3,a)
        editable(ws5,i,4,o); editable(ws5,i,5,s); editable(ws5,i,6)

    fp_ = save_wb(wb,"ClearLedger_Dispute_Resolution_Process_Pack.xlsx")
    return fp_

# =====================================================================
# 2. CREDIT MANAGEMENT
# =====================================================================
def build_credit():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"Credit Management Process Pack - APQC PCF v8.0",font_header14)
    sc(ws,3,2,"Reference: PCF v8.0 Process 10853 - Manage Credit",font_dim)
    header_row(ws,5,["","KPI","Current","Target","Benchmark P50","Benchmark P75","Status"])
    kpis = [("DSO (days)","42","30","35","28","Near"),("Credit check turnaround (hrs)","48","4","8","4","Behind"),("Credit limit utilization (%)","75","85","80","90","Near"),("Past due >90 days (%)","8","3","5","2","Behind"),("Credit hold rate (%)","12","5","8","4","Behind"),("New customer credit setup (days)","7","2","3","1","Behind")]
    for i,k in enumerate(kpis,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,k[0])
        for j,v in enumerate(k[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"Customer Credit")
    col_widths(ws2,[3,25,18,18,18,18,18,18,18,18])
    header_row(ws2,1,["","Customer Name","Credit Limit (EUR)","Used (EUR)","Available (EUR)","Utilization %","Risk Rating","Last Review","Review Due","Status"])
    for r in range(2,16):
        sc(ws2,r,1,"",fill=fill_bg)
        for c in range(2,11): editable(ws2,r,c)
    label(ws2,17,2,"Totals")
    sc(ws2,17,3,"=SUM(C2:C16)",font_editable,fill_input,align_c,nf='#,##0')
    sc(ws2,17,4,"=SUM(D2:D16)",font_editable,fill_input,align_c,nf='#,##0')

    ws3 = new_sheet(wb,"Scoring Matrix")
    col_widths(ws3,[3,30,18,18,18,18])
    header_row(ws3,1,["","Criterion","Weight (%)","Score (1-5)","Weighted","Notes"])
    criteria = [("Payment history","25","","",""),("Financial health","20","","",""),("Industry risk","15","","",""),("Relationship tenure","10","","",""),("Order volume","10","","",""),("Dispute history","10","","",""),("Country risk","10","","","")]
    for i,(c,w,_,_,_) in enumerate(criteria,2):
        sc(ws3,i,1,"",fill=fill_bg); label(ws3,i,2,c); label(ws3,i,3,w)
        editable(ws3,i,4); sc(ws3,i,5,"=C{i}*D{i}/100".replace("{i}",str(i)),font_editable,fill_input,align_c); editable(ws3,i,6)
    label(ws3,10,2,"Total Score"); sc(ws3,10,5,"=SUM(E2:E8)",font_editable,fill_input,align_c)

    ws4 = new_sheet(wb,"Review Schedule")
    col_widths(ws4,[3,25,18,18,18,30])
    header_row(ws4,1,["","Customer","Risk Rating","Last Review","Next Review","Actions Needed"])
    for r in range(2,16):
        sc(ws4,r,1,"",fill=fill_bg)
        for c in range(2,7): editable(ws4,r,c)

    ws5 = new_sheet(wb,"Action Plan")
    col_widths(ws5,[3,8,35,20,20,18])
    header_row(ws5,1,["","Phase","Action","Owner","Target","Status"])
    items = [("1","Audit current credit policies","Credit Manager",""),("1","Implement risk scoring model","Credit/IT",""),("2","Set up automated credit checks","IT/Finance",""),("2","Define credit limit matrix","Credit Manager",""),("3","Deploy periodic review cadence","Credit Team",""),("3","Integrate external bureau data","IT",""),("4","Optimize DSO through targeting","Credit Manager","")]
    for i,(p,a,o,_) in enumerate(items,2):
        sc(ws5,i,1,"",fill=fill_bg); label(ws5,i,2,p); label(ws5,i,3,a)
        editable(ws5,i,4,o); editable(ws5,i,5); editable(ws5,i,6)

    fp_ = save_wb(wb,"ClearLedger_Credit_Management_Process_Pack.xlsx")
    return fp_

# =====================================================================
# 3. COLLECTIONS STRATEGY
# =====================================================================
def build_collections():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"Collections Strategy & Segmentation - APQC PCF v8.0",font_header14)
    sc(ws,3,2,"Reference: PCF v8.0 Process 10854 - Manage Collections",font_dim)
    header_row(ws,5,["","KPI","Current","Target","Benchmark P50","Benchmark P75","Status"])
    kpis = [("CEI (Collection Effectiveness Index %)","75","90","85","95","Behind"),("Average days delinquent","18","8","12","6","Behind"),("Promise-to-pay hit rate (%)","55","75","65","80","Behind"),("Right-party contact rate (%)","60","85","75","90","Behind"),("Roll rate 30->60 days (%)","25","10","15","8","Behind"),("Cost per collection (EUR)","15","6","8","5","Near")]
    for i,k in enumerate(kpis,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,k[0])
        for j,v in enumerate(k[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"Segmentation")
    col_widths(ws2,[3,25,18,18,18,18,18,18,18])
    header_row(ws2,1,["","Segment","Balance (EUR)","% of Total","# Accounts","Strategy","Contact Cadence","Owner","Notes"])
    segs = [("Key / Strategic","","","","VIP treatment","Daily","",""),("High Risk","","","","Intensive","Daily","",""),("Standard","","","","Standard","Weekly","",""),("Low Risk","","","","Light touch","Monthly","",""),("Small balance","","","","Auto-dunning","Auto","","")]
    for i,s in enumerate(segs,2):
        sc(ws2,i,1,"",fill=fill_bg); label(ws2,i,2,s[0])
        for j,v in enumerate(s[1:],3): editable(ws2,i,j,v)
    label(ws2,8,2,"Total"); sc(ws2,8,3,"=SUM(C2:C7)",font_editable,fill_input,align_c)
    sc(ws2,8,4,"=SUM(D2:D7)",font_editable,fill_input,align_c)

    ws3 = new_sheet(wb,"Collection Queue")
    col_widths(ws3,[3,22,18,18,18,18,18,18,18,18])
    header_row(ws3,1,["","Customer","Invoice #","Amount (EUR)","Age (days)","Segment","Assigned To","Last Action","Next Action","Status"])
    for r in range(2,21):
        sc(ws3,r,1,"",fill=fill_bg)
        for c in range(2,11): editable(ws3,r,c)

    ws4 = new_sheet(wb,"Dunning Strategy")
    col_widths(ws4,[3,18,30,18,18,18])
    header_row(ws4,1,["","Day","Action","Channel","Template","Escalation"])
    dun = [(0,"Payment due reminder","Email","Standard",""),(1,"Overdue notice - gentle","Email","D1",""),(7,"Overdue notice - firm","Email + SMS","D7",""),(15,"Overdue notice - final","Email + Letter","D15",""),(30,"Collections call","Phone","","Collector"),(45,"Manager escalation","Phone + Letter","","Credit Manager"),(60,"Credit hold / suspension","System","","Credit Manager"),(90,"Third-party / legal","Legal","","Legal Counsel")]
    for i,(d,a,ch,t,e) in enumerate(dun,2):
        sc(ws4,i,1,"",fill=fill_bg); label(ws4,i,2,str(d)); label(ws4,i,3,a)
        for j,v in enumerate([ch,t,e],4): editable(ws4,i,j,v)

    ws5 = new_sheet(wb,"Action Plan")
    col_widths(ws5,[3,8,35,20,20,18])
    header_row(ws5,1,["","Phase","Action","Owner","Target","Status"])
    items = [("1","Analyze current portfolio aging","AR Lead",""),("1","Define customer segments","Credit/AR",""),("2","Configure dunning rules in ERP","IT/AR",""),("2","Set up collector worklists","AR Lead",""),("3","Deploy automated reminders","IT",""),("3","Implement promise-to-pay tracking","AR",""),("4","Monthly CEI review and optimization","AR Manager","")]
    for i,(p,a,o,_) in enumerate(items,2):
        sc(ws5,i,1,"",fill=fill_bg); label(ws5,i,2,p); label(ws5,i,3,a)
        editable(ws5,i,4,o); editable(ws5,i,5); editable(ws5,i,6)

    fp_ = save_wb(wb,"ClearLedger_Collections_Strategy_Segmentation.xlsx")
    return fp_

# =====================================================================
# 4. BILLING & INVOICING
# =====================================================================
def build_billing():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"Billing & Invoicing Process Pack - APQC PCF v8.0",font_header14)
    sc(ws,3,2,"Reference: PCF v8.0 Process 10851 - Bill Customers",font_dim)
    header_row(ws,5,["","KPI","Current","Target","Benchmark P50","Benchmark P75","Status"])
    kpis = [("Invoice accuracy rate (%)","92","99","97","99.5","Behind"),("First-pass billing rate (%)","78","95","90","97","Behind"),("Time to invoice (days)","5","1","2","1","Behind"),("E-invoice adoption (%)","45","80","65","90","Behind"),("Billing error rate (%)","8","1","3","0.5","Behind"),("Cost per invoice (EUR)","2.50","1.00","1.50","0.80","Near")]
    for i,k in enumerate(kpis,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,k[0])
        for j,v in enumerate(k[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"Invoice Types")
    col_widths(ws2,[3,30,18,18,18,18,18])
    header_row(ws2,1,["","Invoice Type","Volume/Month","Avg Value (EUR)","Format","Delivery Method","Automation Status"])
    types = [("Standard invoice","","","PDF/EDI","Email",""),("Credit memo","","","PDF","Email",""),("Debit memo","","","PDF","Email",""),("Proforma invoice","","","PDF","Email",""),("Consolidated invoice","","","PDF/EDI","Portal",""),("E-invoice (PEPPOL)","","","XML","PEPPOL",""),("Self-billed invoice","","","XML","EDI",""),("Recurring invoice","","","PDF","Auto","")]
    for i,t in enumerate(types,2):
        sc(ws2,i,1,"",fill=fill_bg); label(ws2,i,2,t[0])
        for j,v in enumerate(t[1:],3): editable(ws2,i,j,v)

    ws3 = new_sheet(wb,"Billing Schedule")
    col_widths(ws3,[3,25,18,18,18,18,30])
    header_row(ws3,1,["","Customer","Billing Cycle","Invoice Day","Cut-off Date","E-invoice?","Notes"])
    for r in range(2,21):
        sc(ws3,r,1,"",fill=fill_bg)
        for c in range(2,8): editable(ws3,r,c)

    ws4 = new_sheet(wb,"Error Tracking")
    col_widths(ws4,[3,22,18,18,18,18,18,18])
    header_row(ws4,1,["","Date","Customer","Invoice #","Error Type","Root Cause","Resolution","Status"])
    for r in range(2,21): sc(ws4,r,1,"",fill=fill_bg); [editable(ws4,r,c) for c in range(2,9)]

    ws5 = new_sheet(wb,"Compliance Matrix")
    col_widths(ws5,[3,25,18,18,18,30])
    header_row(ws5,1,["","Country","E-invoice Mandate","Mandate Date","Compliant?","Notes"])
    countries = [("France","Y","2024","",""),("Germany","Y","2025","",""),("Italy","Y","2019","",""),("Spain","Y","2024","",""),("Poland","Y","2024","",""),("Romania","Y","2024","",""),("Belgium","Y","2026","",""),("Portugal","Y","2025","",""),("Netherlands","N","N/A","",""),("UK","N","N/A","","")]
    for i,c in enumerate(countries,2):
        sc(ws5,i,1,"",fill=fill_bg); label(ws5,i,2,c[0])
        for j,v in enumerate(c[1:],3): editable(ws5,i,j,v)

    fp_ = save_wb(wb,"ClearLedger_Billing_Invoicing_Process_Pack.xlsx")
    return fp_

# =====================================================================
# 5. SOX CONTROLS
# =====================================================================
def build_sox():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"SOX Controls Library - APQC PCF v8.0",font_header14)
    sc(ws,3,2,"Reference: COSO 2013 / SOX Section 404 Compliance",font_dim)
    header_row(ws,5,["","Control Area","# Controls","Implemented","Tested","Pass Rate","Status"])
    areas = [("Cash Application","12","8","6","88","In Progress"),("Credit Management","8","6","5","80","In Progress"),("Collections","10","7","5","86","In Progress"),("Dispute Resolution","8","5","4","75","Needs Work"),("Billing & Invoicing","15","12","10","92","On Track"),("Period-end Close","10","9","8","90","On Track"),("System Access / Segregation","14","12","11","95","On Track"),("Third-party / Vendor","6","4","3","83","In Progress")]
    for i,a in enumerate(areas,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,a[0])
        for j,v in enumerate(a[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"Control Matrix")
    col_widths(ws2,[3,8,30,25,18,18,18,18])
    header_row(ws2,1,["","ID","Control Description","Process Area","Control Type","Frequency","Owner","Status"])
    controls = [
        ("CA-01","Automated matching of remittance to open invoices","Cash App","Automated","Daily","",""),
        ("CA-02","Review of unapplied cash aging report","Cash App","Detective","Weekly","",""),
        ("CA-03","Segregation: cash app vs. adjustments","Cash App","Preventive","Quarterly","",""),
        ("CM-01","Credit limit approval matrix","Credit","Preventive","Per request","",""),
        ("CM-02","Credit scoring model validation","Credit","Detective","Monthly","",""),
        ("CL-01","Collection call logging and review","Collections","Detective","Weekly","",""),
        ("CL-02","Dunning process automated triggers","Collections","Automated","Daily","",""),
        ("DR-01","Dispute logging and categorization","Disputes","Preventive","Daily","",""),
        ("DR-02","Write-off approval authorization","Disputes","Preventive","Per request","",""),
        ("BI-01","Invoice auto-generation from contracts","Billing","Automated","Daily","",""),
        ("BI-02","Billing error rate monitoring","Billing","Detective","Monthly","",""),
        ("SA-01","User access review - AR systems","Access","Detective","Quarterly","",""),
        ("SA-02","Segregation of duties matrix","Access","Preventive","Annually","",""),
        ("PC-01","AR sub-ledger to GL reconciliation","Close","Detective","Monthly","",""),
    ]
    for i,ctrl in enumerate(controls,2):
        sc(ws2,i,1,"",fill=fill_bg); label(ws2,i,2,ctrl[0]); label(ws2,i,3,ctrl[1])
        for j,v in enumerate(ctrl[2:],4): editable(ws2,i,j,v)

    ws3 = new_sheet(wb,"Testing Log")
    col_widths(ws3,[3,18,22,18,18,18,18,18,18])
    header_row(ws3,1,["","Control ID","Tester","Test Date","Result (Pass/Fail)","Deficiency","Remediation Plan","Due Date","Status"])
    for r in range(2,31): sc(ws3,r,1,"",fill=fill_bg); [editable(ws3,r,c) for c in range(2,10)]

    ws4 = new_sheet(wb,"Action Plan")
    col_widths(ws4,[3,30,18,18,18,30])
    header_row(ws4,1,["","Control Gap","Risk Level","Owner","Target Date","Remediation Plan"])
    for r in range(2,21): sc(ws4,r,1,"",fill=fill_bg); [editable(ws4,r,c) for c in range(2,7)]

    fp_ = save_wb(wb,"ClearLedger_SOX_Controls_Library.xlsx")
    return fp_

# =====================================================================
# 6. TECH SELECTION GUIDE
# =====================================================================
def build_tech_select():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,20,20,20,20])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"Technology Selection Guide - AR Automation",font_header14)
    sc(ws,3,2,"Framework for evaluating and selecting AR technology solutions",font_dim)
    header_row(ws,5,["","Evaluation Criteria","Weight (%)","Score (1-5)","Weighted Score","Notes"])
    criteria = [
        ("Auto-cash matching engine","20","","",""),
        ("ERP integration depth","15","","",""),
        ("E-invoicing / PEPPOL compliance","15","","",""),
        ("AI / ML capabilities","10","","",""),
        ("Scalability & performance","10","","",""),
        ("User experience & adoption","10","","",""),
        ("Total cost of ownership","10","","",""),
        ("Vendor support & SLAs","5","","",""),
        ("Security & certifications","5","","",""),
    ]
    for i,(c,w,_,_,_) in enumerate(criteria,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,c); label(ws,i,3,w)
        editable(ws,i,4); sc(ws,i,5,f"=C{i}*D{i}/100",font_editable,fill_input,align_c); editable(ws,i,6)
    label(ws,16,2,"Total Weighted Score"); sc(ws,16,5,"=SUM(E6:E14)",font_editable,fill_input,align_c)

    ws2 = new_sheet(wb,"Vendor Comparison")
    col_widths(ws2,[3,25,18,18,18,18,18,18])
    header_row(ws2,1,["","Vendor","Product","Auto-match","E-invoicing","ERP Integration","Pricing Model","Score"])
    vendors = [("HighRadius","AR Cloud","","","","",""),("Billtrust","AR Suite","","","","",""),("Versapay","AR Automation","","","","",""),("Kofax","TotalAgility","","","","",""),("Esker","AR Platform","","","","",""),("Bottomline","PayStream","","","","",""),("Coupa","AR Module","","","","",""),("ClearLedger","OtC Suite","","","","","")]
    for i,v in enumerate(vendors,2):
        sc(ws2,i,1,"",fill=fill_bg); label(ws2,i,2,v[0]); label(ws2,i,3,v[1])
        for j,val in enumerate(v[2:],4): editable(ws2,i,j,val)

    ws3 = new_sheet(wb,"Requirements Checklist")
    col_widths(ws3,[3,40,18,18,30])
    header_row(ws3,1,["","Requirement","Priority (H/M/L)","Required? (Y/N)","Notes"])
    reqs = [("Automated cash application","H","Y",""),("Multi-ERP support","H","Y",""),("E-invoice PEPPOL sending/receiving","H","Y",""),("Credit card payment processing","M","",""),("Customer portal for invoice presentment","M","",""),("Deduction / dispute workbench","H","Y",""),("Collections worklist & automation","H","Y",""),("Real-time reporting & dashboards","M","",""),("AI-powered remittance prediction","M","",""),("Role-based access control","H","Y",""),("API for custom integrations","M","",""),("Mobile access for approvals","L","","")]
    for i,r in enumerate(reqs,2):
        sc(ws3,i,1,"",fill=fill_bg); label(ws3,i,2,r[0]); label(ws3,i,3,r[1])
        editable(ws3,i,4,r[2]); editable(ws3,i,5)

    ws4 = new_sheet(wb,"ROI Comparison")
    col_widths(ws4,[3,30,18,18,18,18])
    header_row(ws4,1,["","Vendor","Impl. Cost (EUR)","Annual Fee (EUR)","Est. Savings/Yr (EUR)","Payback (months)"])
    for r in range(2,12):
        sc(ws4,r,1,"",fill=fill_bg)
        for c in range(2,7): editable(ws4,r,c)

    fp_ = save_wb(wb,"ClearLedger_Technology_Selection_Guide.xlsx")
    return fp_

# =====================================================================
# 7. PROCESS MINING PLAYBOOK
# =====================================================================
def build_mining():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"Process Mining Playbook - APQC PCF v8.0",font_header14)
    sc(ws,3,2,"Framework for mining, analyzing, and optimizing OtC processes",font_dim)
    header_row(ws,5,["","Mining Area","Current State","Target State","Gap","Priority","Timeline"])
    areas = [("Cash application flow","Manual / hybrid","80%+ auto-match","","H","Q1"),("Invoice-to-cash cycle","14 days","5 days","","H","Q1"),("Dispute resolution path","9 steps","5 steps","","M","Q2"),("Credit check process","48 hrs","4 hrs","","M","Q2"),("Collection touch pattern","Reactive","Segmented","","H","Q1"),("Dunning automation","Manual","Fully auto","","M","Q2"),("Remittance matching","65%","95%","","H","Q1")]
    for i,a in enumerate(areas,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,a[0])
        for j,v in enumerate(a[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"Event Log Template")
    col_widths(ws2,[3,18,22,22,18,18,18,18])
    header_row(ws2,1,["","Case ID","Activity","Timestamp","Resource","Step Duration (hrs)","Wait Time (hrs)","Cost (EUR)"])
    for r in range(2,26): sc(ws2,r,1,"",fill=fill_bg); [editable(ws2,r,c) for c in range(2,9)]

    ws3 = new_sheet(wb,"Variant Analysis")
    col_widths(ws3,[3,8,35,18,18,18])
    header_row(ws3,1,["","Variant","Path Description","Frequency (%)","Avg Cycle Time","Cost Impact"])
    for r in range(2,16): sc(ws3,r,1,"",fill=fill_bg); [editable(ws3,r,c) for c in range(2,7)]

    ws4 = new_sheet(wb,"Bottleneck Identification")
    col_widths(ws4,[3,30,18,18,18,18,18])
    header_row(ws4,1,["","Process Step","Avg Wait (hrs)","Max Wait (hrs)","Volume","Root Cause","Recommendation"])
    for r in range(2,16): sc(ws4,r,1,"",fill=fill_bg); [editable(ws4,r,c) for c in range(2,8)]

    fp_ = save_wb(wb,"ClearLedger_Process_Mining_Playbook.xlsx")
    return fp_

# =====================================================================
# 8. SSC TRANSITION GUIDE
# =====================================================================
def build_ssc():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"SSC Transition Guide - Shared Services AR Setup",font_header14)
    sc(ws,3,2,"End-to-end framework for transitioning AR to a shared service center",font_dim)
    header_row(ws,5,["","Phase","Duration (weeks)","Resources Needed","Key Deliverable","Status","Risk Level"])
    phases = [("1 - Assess & Scope","4","","Transition charter","","M"),("2 - Design Target Ops Model","6","","To-be process maps","","M"),("3 - Legal & Entity Setup","8","","Entity registration","","H"),("4 - System & ERP Config","10","","Configured systems","","H"),("5 - Process Migration","12","","Migrated processes","","H"),("6 - People & Training","6","","Trained resources","","M"),("7 - Shadow Run & Stabilize","8","","Parallel run sign-off","","H"),("8 - Go-Live & Hypercare","4","","Go-live certificate","","H")]
    for i,p in enumerate(phases,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,p[0])
        for j,v in enumerate(p[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"Transition Plan")
    col_widths(ws2,[3,8,35,20,18,18,18])
    header_row(ws2,1,["","Step","Activity","Owner","Start","End","Status"])
    for r in range(2,41): sc(ws2,r,1,"",fill=fill_bg); [editable(ws2,r,c) for c in range(2,8)]

    ws3 = new_sheet(wb,"SLA Framework")
    col_widths(ws3,[3,30,18,18,18,18])
    header_row(ws3,1,["","Process","KPI","Target","Measurement","Frequency"])
    slas = [("Cash application","Auto-cash rate","85%","System report","Daily"),("Dispute resolution","Resolution time","5 days","Case system","Weekly"),("Collections","CEI","90%","System calc","Monthly"),("Credit management","Credit check TAT","4 hrs","System log","Weekly"),("Billing","Invoice accuracy","99%","Audit sample","Monthly"),("Customer inquiry","Response time","4 hrs","CRM","Weekly"),("Period-end close","Close time","3 days","Actual","Monthly")]
    for i,s in enumerate(slas,2):
        sc(ws3,i,1,"",fill=fill_bg); label(ws3,i,2,s[0]); label(ws3,i,3,s[1])
        for j,v in enumerate(s[2:],4): editable(ws3,i,j,v)

    ws4 = new_sheet(wb,"Cost Model")
    col_widths(ws4,[3,35,18,18,18,18])
    header_row(ws4,1,["","Cost Category","Current (EUR)","Target (EUR)","Savings (EUR)","Notes"])
    costs = [("FTE cost - AR","","","",""),("Technology / licenses","","","",""),("Infrastructure","","","",""),("Training & change mgmt","","","",""),("Transition one-time costs","","","",""),("Ongoing operations","","","",""),("Total","=SUM(B2:B7)","=SUM(C2:C7)","=SUM(D2:D7)","")]
    for i,c in enumerate(costs,2):
        sc(ws4,i,1,"",fill=fill_bg); label(ws4,i,2,c[0])
        for j,v in enumerate(c[1:],3):
            if v: sc(ws4,i,j,v,font_editable,fill_input,align_c)
            else: editable(ws4,i,j)

    fp_ = save_wb(wb,"ClearLedger_SSC_Transition_Guide.xlsx")
    return fp_

# =====================================================================
# 9. TREASURY & WC GUIDE
# =====================================================================
def build_treasury():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"Treasury & Working Capital Guide - APQC PCF v8.0",font_header14)
    sc(ws,3,2,"Reference: PCF v8.0 Process 10855 - Manage Treasury Operations",font_dim)
    header_row(ws,5,["","Metric","Current","Target","Benchmark P50","Benchmark P75","Status"])
    kpis = [("DSO","45","30","35","28","Near"),("DIO","55","45","50","40","Near"),("DPO","38","45","42","50","Near"),("CCC (days)","62","30","43","18","Behind"),("Cash conversion efficiency","65","85","78","90","Behind"),("Working capital / Revenue (%)","18","12","15","10","Near"),("Forecast accuracy (%)","75","90","85","95","Behind")]
    for i,k in enumerate(kpis,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,k[0])
        for j,v in enumerate(k[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"WC Calculator")
    col_widths(ws2,[3,35,20,20,20])
    header_row(ws2,1,["","Component","Value (EUR)","Target (EUR)","Variance"])
    wc_items = [("Accounts Receivable","","",""),("Accounts Payable","","",""),("Inventory","","",""),("Net Working Capital","=B2-B3+B4","=C2-C3+C4","=D2-D3+D4")]
    for i,w in enumerate(wc_items,2):
        sc(ws2,i,1,"",fill=fill_bg); label(ws2,i,2,w[0])
        for j,v in enumerate(w[1:],3):
            if "=" in str(v): sc(ws2,i,j,v,font_editable,fill_input,align_c,nf='#,##0')
            else: editable(ws2,i,j,v,nf='#,##0')

    ws3 = new_sheet(wb,"Cash Forecasting")
    col_widths(ws3,[3,18,18,18,18,18,18,18])
    header_row(ws3,1,["","Week 1","Week 2","Week 3","Week 4","Month +1","Month +2","Month +3"])
    for r in range(2,12): sc(ws3,r,1,"",fill=fill_bg); [editable(ws3,r,c) for c in range(2,9)]

    ws4 = new_sheet(wb,"Bank Relationship")
    col_widths(ws4,[3,25,18,18,30,18])
    header_row(ws4,1,["","Bank","Account Type","Currency","Account #","Contact"])
    for r in range(2,12): sc(ws4,r,1,"",fill=fill_bg); [editable(ws4,r,c) for c in range(2,7)]

    ws5 = new_sheet(wb,"Action Plan")
    col_widths(ws5,[3,8,35,20,20,18])
    header_row(ws5,1,["","Phase","Action","Owner","Target","Status"])
    items = [("1","Calculate current CCC baseline","Treasury",""),("1","Identify DSO reduction levers","AR/Treasury",""),("2","Optimize payment terms structure","Treasury",""),("2","Implement cash forecasting tool","IT/Treasury",""),("3","Set up working capital targets","CFO/Treasury",""),("3","Monthly WC review with business","Treasury",""),("4","Benchmark against APQC P50+","Treasury","")]
    for i,(p,a,o,_) in enumerate(items,2):
        sc(ws5,i,1,"",fill=fill_bg); label(ws5,i,2,p); label(ws5,i,3,a)
        editable(ws5,i,4,o); editable(ws5,i,5); editable(ws5,i,6)

    fp_ = save_wb(wb,"ClearLedger_Treasury_Working_Capital_Guide.xlsx")
    return fp_

# =====================================================================
# 10. CUSTOMER ONBOARDING
# =====================================================================
def build_onboarding():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = HEADER_BLUE
    col_widths(ws,[3,35,18,18,18,18,18])
    sc(ws,1,2,"ClearLedger",Font(name="Aptos",size=20,bold=True,color=HEADER_BLUE))
    sc(ws,2,2,"Customer Onboarding Playbook - APQC PCF v8.0",font_header14)
    sc(ws,3,2,"Streamlined AR setup for new customers, aligned with 2026 benchmarks",font_dim)
    header_row(ws,5,["","KPI","Current","Target","Benchmark","Status"])
    kpis = [("New customer setup time (days)","7","2","2","Behind"),("Credit check turnaround (hrs)","48","4","4","Behind"),("First invoice accuracy (%)","85","99","99","Behind"),("Time to first payment (days)","45","30","30","Near"),("Customer info completeness (%)","70","100","100","Behind"),("Portal adoption rate (%)","40","80","75","Behind")]
    for i,k in enumerate(kpis,6):
        sc(ws,i,1,"",fill=fill_bg); label(ws,i,2,k[0])
        for j,v in enumerate(k[1:],3): editable(ws,i,j,v)

    ws2 = new_sheet(wb,"Onboarding Workflow")
    col_widths(ws2,[3,8,35,20,18,18,18])
    header_row(ws2,1,["","Step","Activity","Owner","SLA (hrs)","Sequence","Status"])
    steps = [("1","Customer information collection","Sales","2","",""),("2","Credit check & scoring","Credit","4","",""),("3","Credit limit approval","Credit Manager","8","",""),("4","Customer master data setup","Master Data","4","",""),("5","Terms & conditions entry","AR","2","",""),("6","ERP customer creation","IT/AR","4","",""),("7","Bank account validation","Treasury","2","",""),("8","E-invoice / PEPPOL setup","IT","8","",""),("9","Portal account creation","IT","2","",""),("10","Welcome email & training","AR/AM","2","",""),("11","First invoice generation","AR","4","",""),("12","Confirmation call","AM","1","","")]
    for i,s in enumerate(steps,2):
        sc(ws2,i,1,"",fill=fill_bg); label(ws2,i,2,s[0]); label(ws2,i,3,s[1])
        for j,v in enumerate(s[2:],4): editable(ws2,i,j,v)

    ws3 = new_sheet(wb,"New Customer Tracker")
    col_widths(ws3,[3,22,18,18,18,18,18,18])
    header_row(ws3,1,["","Company","Contact","Credit Limit","Terms","Portal Access","Onboarding Date","Status"])
    for r in range(2,21): sc(ws3,r,1,"",fill=fill_bg); [editable(ws3,r,c) for c in range(2,9)]

    ws4 = new_sheet(wb,"Document Checklist")
    col_widths(ws4,[3,35,18,18,18])
    header_row(ws4,1,["","Document","Required (Y/N)","Received","Notes"])
    docs = [("Signed customer agreement","Y","",""),("Tax registration certificate","Y","",""),("Bank account confirmation","Y","",""),("Credit application form","Y","",""),("Financial statements (if requested)","M","",""),("PEPPOL / E-invoicing ID","Y","",""),("PO / contract reference","Y","",""),("Vendor registration form","Y","",""),("Contact list (AR/payables)","Y","","")]
    for i,d in enumerate(docs,2):
        sc(ws4,i,1,"",fill=fill_bg); label(ws4,i,2,d[0]); label(ws4,i,3,d[1])
        editable(ws4,i,4,""); editable(ws4,i,5)

    ws5 = new_sheet(wb,"Action Plan")
    col_widths(ws5,[3,8,35,20,20,18])
    header_row(ws5,1,["","Phase","Action","Owner","Target","Status"])
    items = [("1","Document current onboarding process","AR Lead",""),("1","Define credit decision matrix","Credit",""),("2","Automate customer setup in ERP","IT",""),("2","Create welcome package & training","AM/AR",""),("3","Implement portal auto-provisioning","IT",""),("3","Set up KPI tracking dashboard","AR/IT",""),("4","Quarterly process review","AR Manager","")]
    for i,(p,a,o,_) in enumerate(items,2):
        sc(ws5,i,1,"",fill=fill_bg); label(ws5,i,2,p); label(ws5,i,3,a)
        editable(ws5,i,4,o); editable(ws5,i,5); editable(ws5,i,6)

    fp_ = save_wb(wb,"ClearLedger_Customer_Onboarding_Playbook.xlsx")
    return fp_

# =====================================================================
# MAIN
# =====================================================================
if __name__ == "__main__":
    builders = [
        ("Dispute Resolution", build_dispute),
        ("Credit Management", build_credit),
        ("Collections Strategy", build_collections),
        ("Billing & Invoicing", build_billing),
        ("SOX Controls", build_sox),
        ("Tech Selection Guide", build_tech_select),
        ("Process Mining Playbook", build_mining),
        ("SSC Transition Guide", build_ssc),
        ("Treasury & WC Guide", build_treasury),
        ("Customer Onboarding", build_onboarding),
    ]
    results = []
    for name, builder in builders:
        if builder:
            try:
                fp = builder()
                results.append(f"OK: {name} -> {fp}")
                print(f"OK: {name}")
            except Exception as e:
                results.append(f"FAIL: {name} - {e}")
                print(f"FAIL: {name} - {e}")
    print("---")
    for r in results:
        print(r)
