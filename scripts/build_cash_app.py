import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

wb = openpyxl.Workbook()

# ── color palette ──
BG = "0D1117"
SURFACE = "161B22"
INPUT_BG = "1A2332"
HEADER_BLUE = "58A6FF"
EDIT_BLUE = "4488FF"
LABEL = "E6EDF3"
DIM = "8B949E"
NOTE = "484F58"
WHITE = "FFFFFF"
RED_ACCENT = "F85149"
GREEN_ACCENT = "3FB950"

font_body = Font(name="Aptos", size=11, color=LABEL)
font_header = Font(name="Aptos", size=14, bold=True, color=HEADER_BLUE)
font_section = Font(name="Aptos", size=12, bold=True, color=HEADER_BLUE)
font_editable = Font(name="Aptos", size=11, color=EDIT_BLUE)
font_dim = Font(name="Aptos", size=10, color=DIM)
font_note = Font(name="Aptos", size=10, color=NOTE)
font_white_bold = Font(name="Aptos", size=11, bold=True, color=WHITE)
font_title = Font(name="Aptos", size=16, bold=True, color=HEADER_BLUE)

fill_bg = PatternFill(start_color=BG, end_color=BG, fill_type="solid")
fill_surface = PatternFill(start_color=SURFACE, end_color=SURFACE, fill_type="solid")
fill_input = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
fill_header = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
fill_green = PatternFill(start_color=GREEN_ACCENT, end_color=GREEN_ACCENT, fill_type="solid")
fill_red = PatternFill(start_color=RED_ACCENT, end_color=RED_ACCENT, fill_type="solid")

thin_border = Border(
    left=Side(style="thin",color="30363D"),
    right=Side(style="thin",color="30363D"),
    top=Side(style="thin",color="30363D"),
    bottom=Side(style="thin",color="30363D"),
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_cell(ws, row, col, value, font=font_body, fill=fill_bg, alignment=align_left, border=thin_border, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell

def style_range(ws, row, col_start, col_end, values, font=font_body, fill=fill_bg, alignment=align_left):
    for i, v in enumerate(values):
        style_cell(ws, row, col_start + i, v, font, fill, alignment)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def make_editable(ws, row, col, value, number_format=None):
    style_cell(ws, row, col, value, font_editable, fill_input, align_center, number_format=number_format)

def make_label(ws, row, col, value):
    style_cell(ws, row, col, value, font_body, fill_surface)

def make_section_header(ws, row, col, value, ncols=1):
    for c in range(col, col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = font_section
        cell.fill = PatternFill(start_color=SURFACE, end_color=SURFACE, fill_type="solid")
        cell.alignment = align_left
        cell.border = thin_border
    ws.cell(row=row, column=col, value=value)

def make_header_row(ws, row, values):
    for i, v in enumerate(values, 1):
        style_cell(ws, row, i, v, font_white_bold, fill_header, align_center)

# =====================================================================
# SHEET 1: Overview
# =====================================================================
ws_ov = wb.active
ws_ov.title = "Overview"
ws_ov.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_ov, [3, 35, 20, 20, 20, 20, 20, 20])

style_cell(ws_ov, 2, 2, "ClearLedger", Font(name="Aptos", size=20, bold=True, color=HEADER_BLUE), fill_bg)
style_cell(ws_ov, 3, 2, "Cash Application Process Pack — APQC PCF v8.0", font_title, fill_bg)
style_cell(ws_ov, 4, 2, "Reference: PCF v8.0 Process 10855 — Apply Cash", font_dim, fill_bg)

headers = ["", "Metric", "Current Value", "Target", "Benchmark P50", "Benchmark P75", "Status", "Priority"]
for i, h in enumerate(headers, 1):
    if i == 1:
        style_cell(ws_ov, 6, i, h, font_white_bold, fill_header, align_left)
    else:
        style_cell(ws_ov, 6, i, h, font_white_bold, fill_header, align_center)

metrics = [
    ("Auto-cash rate (%)", "60", "85", "78", "92", "Behind", "High"),
    ("Straight-through processing (%)", "55", "80", "72", "88", "Behind", "High"),
    ("Days to apply (DTA)", "2.5", "1.0", "1.2", "0.8", "Near", "Medium"),
    ("Unapplied cash (%)", "8", "3", "4", "2", "Behind", "High"),
    ("Deduction rate (%)", "4.5", "2.0", "2.8", "1.5", "Near", "Medium"),
    ("Match rate remittance (%)", "65", "90", "85", "95", "Behind", "High"),
    ("Cost per invoice processed (EUR)", "3.20", "1.50", "1.80", "1.20", "Near", "Medium"),
    ("ERP sync lag (hours)", "4", "1", "2", "1", "Near", "High"),
]

for r, m in enumerate(metrics, 7):
    style_cell(ws_ov, r, 1, "", fill_bg)
    make_label(ws_ov, r, 2, m[0])
    for c, v in enumerate(m[1:], 3):
        make_editable(ws_ov, r, c, v)

style_cell(ws_ov, 16, 2, "Instructions: Update Current Value columns with your data. Targets and benchmarks are APQC PCF v8.0 2026.", font_note, fill_bg)

# =====================================================================
# SHEET 2: Data Entry
# =====================================================================
ws_de = wb.create_sheet("Data Entry")
ws_de.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_de, [3, 30, 18, 18, 18, 18, 18])

make_header_row(ws_de, 1, ["", "Entity / Process Step", "Volume (monthly)", "FTE Allocation", "Avg Hours/Unit", "Error Rate (%)", "Notes"])

entities = [
    "Lockbox processing",
    "EDI/810 remittance",
    "Check payments",
    "ACH/Wire receipts",
    "Credit card payments",
    "Portal remittance retrieval",
    "Manual cash application",
    "Deduction coding",
    "Unapplied cash research",
    "ERP posting & reconciliation",
]
for r, e in enumerate(entities, 2):
    style_cell(ws_de, r, 1, "", fill_bg)
    make_label(ws_de, r, 2, e)
    for c in range(3, 7):
        make_editable(ws_de, r, c, "")

style_cell(ws_de, 13, 2, "Total monthly invoice volume processed:", font_body, fill_surface)
ws_de.cell(row=13, column=3).value = "=SUM(C2:C11)"
ws_de.cell(row=13, column=3).font = Font(name="Aptos", size=11, bold=True, color=EDIT_BLUE)
ws_de.cell(row=13, column=3).fill = fill_input
ws_de.cell(row=13, column=3).alignment = align_center
ws_de.cell(row=13, column=3).border = thin_border

for c_letter in ["D", "E", "F", "G"]:
    cell = ws_de[f"{c_letter}13"]
    cell.value = f"=SUM({c_letter}2:{c_letter}11)"
    cell.font = Font(name="Aptos", size=11, bold=True, color=EDIT_BLUE)
    cell.fill = fill_input
    cell.alignment = align_center
    cell.border = thin_border

style_cell(ws_de, 15, 2, "Auto-cash match rules configured? Enter Y/N:", font_body, fill_surface)
make_editable(ws_de, 15, 3, "")

style_cell(ws_de, 16, 2, "ERP integration active? Enter Y/N:", font_body, fill_surface)
make_editable(ws_de, 16, 3, "")

# =====================================================================
# SHEET 3: Remittance Matching
# =====================================================================
ws_rm = wb.create_sheet("Remittance Matching")
ws_rm.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_rm, [3, 30, 18, 18, 18, 18])

make_header_row(ws_rm, 1, ["", "Remittance Source", "Active? (Y/N)", "Match Rate (%)", "Avg Delay (hrs)", "Automation Status"])

sources = [
    "EDI 820",
    "Check remittance stub",
    "ACH/Wire details",
    "Customer portal download",
    "Email PDF parsing",
    "Supplier portal (manual)",
    "ERP auto-match engine",
    "Third-party matching tool",
]
for r, s in enumerate(sources, 2):
    style_cell(ws_rm, r, 1, "", fill_bg)
    make_label(ws_rm, r, 2, s)
    for c in range(3, 7):
        make_editable(ws_rm, r, c, "")

# =====================================================================
# SHEET 4: Deduction Tracking
# =====================================================================
ws_dt = wb.create_sheet("Deduction Tracking")
ws_dt.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_dt, [3, 20, 18, 18, 18, 18, 18, 18, 18])

make_section_header(ws_dt, 1, 2, "Deduction Register — Enter new items below", 8)
make_header_row(ws_dt, 2, ["", "Customer", "Invoice #", "Deduction Amt (€)", "Reason Code", "Date Opened", "Status", "Resolution Date", "Notes"])

for r in range(3, 13):
    style_cell(ws_dt, r, 1, "", fill_bg)
    for c in range(2, 10):
        make_editable(ws_dt, r, c, "")

style_cell(ws_dt, 14, 2, "Total deductions:", font_body, fill_surface)
cell_sum_d = ws_dt.cell(row=14, column=4, value="=SUM(D3:D12)")
cell_sum_d.font = Font(name="Aptos", size=11, bold=True, color=EDIT_BLUE)
cell_sum_d.fill = fill_input
cell_sum_d.alignment = align_center
cell_sum_d.border = thin_border

# =====================================================================
# SHEET 5: Aging Analysis
# =====================================================================
ws_ag = wb.create_sheet("Aging Analysis")
ws_ag.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_ag, [3, 25, 18, 18, 18, 18, 18])

make_header_row(ws_ag, 1, ["", "Bucket", "Balance (€)", "% of Total", "Target %", "Variance", "Action Priority"])

buckets = [
    "Current (0–30 days)",
    "31–60 days",
    "61–90 days",
    "91–120 days",
    "Over 120 days",
]
for r, b in enumerate(buckets, 2):
    style_cell(ws_ag, r, 1, "", fill_bg)
    make_label(ws_ag, r, 2, b)
    for c in [3, 4]:
        make_editable(ws_ag, r, c, "")
    make_editable(ws_ag, r, 5, "")
    ws_ag.cell(row=r, column=6).value = f"=D{r}-E{r}"
    ws_ag.cell(row=r, column=6).font = Font(name="Aptos", size=11, bold=True, color=EDIT_BLUE)
    ws_ag.cell(row=r, column=6).fill = fill_input
    ws_ag.cell(row=r, column=6).alignment = align_center
    ws_ag.cell(row=r, column=6).border = thin_border
    make_editable(ws_ag, r, 7, "")

style_cell(ws_ag, 8, 2, "Total:", Font(name="Aptos", size=11, bold=True, color=LABEL), fill_surface)
for c_letter in ["C", "D", "E", "F"]:
    cell = ws_ag[f"{c_letter}8"]
    cell.value = f"=SUM({c_letter}2:{c_letter}7)"
    cell.font = Font(name="Aptos", size=11, bold=True, color=EDIT_BLUE)
    cell.fill = fill_input
    cell.alignment = align_center
    cell.border = thin_border

# =====================================================================
# SHEET 6: KPIs & Dashboard
# =====================================================================
ws_kpi = wb.create_sheet("KPIs & Dashboard")
ws_kpi.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_kpi, [3, 35, 20, 20, 20, 20])

make_header_row(ws_kpi, 1, ["", "KPI", "Formula / Source", "Current", "Target", "Status"])

kpis = [
    ("Auto-cash rate", "=Data Entry!C13", "", "85%", ""),
    ("STP rate", "Remittance Matching match rate avg", "", "80%", ""),
    ("Days to apply", "Manual entry", "", "1.0", ""),
    ("Unapplied cash %", "=Unapplied Cash!C7", "", "3%", ""),
    ("Cost per invoice", "=Data Entry!C13/(Data Entry!D13*160)", "", "€1.50", ""),
    ("Deduction rate", "=Deduction Tracking!D14/Data Entry!C13*100", "", "2.0%", ""),
    ("ERP sync lag (hrs)", "Manual entry", "", "1", ""),
]

for r, k in enumerate(kpis, 2):
    style_cell(ws_kpi, r, 1, "", fill_bg)
    make_label(ws_kpi, r, 2, k[0])
    make_label(ws_kpi, r, 3, k[1])
    for c in range(4, 6):
        make_editable(ws_kpi, r, c, k[c-2])
    style_cell(ws_kpi, r, 6, k[4] if len(k) > 4 else "")

# =====================================================================
# SHEET 7: Automation Calculator
# =====================================================================
ws_ac = wb.create_sheet("Automation Calculator")
ws_ac.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_ac, [3, 35, 20, 20, 20, 20])

make_section_header(ws_ac, 1, 2, "ROI Calculator — Enter your values", 5)
make_header_row(ws_ac, 2, ["", "Parameter", "Current", "Target", "Unit", "Notes"])

roi_params = [
    ("Monthly invoice volume", "", "", "", "From Data Entry tab"),
    ("FTE cost (fully loaded)", "4500", "", "€/month", "Per FTE"),
    ("Current FTEs on cash app", "", "", "headcount", ""),
    ("Target FTEs after automation", "", "", "headcount", ""),
    ("Current error rate", "", "", "%", ""),
    ("Target error rate", "", "", "%", ""),
    ("Implementation cost (one-time)", "25000", "", "€", "Software + services"),
    ("Annual maintenance", "5000", "", "€/year", ""),
]

for r, p in enumerate(roi_params, 3):
    style_cell(ws_ac, r, 1, "", fill_bg)
    make_label(ws_ac, r, 2, p[0])
    make_editable(ws_ac, r, 3, p[1])
    make_editable(ws_ac, r, 4, p[2])
    make_label(ws_ac, r, 5, p[3])
    make_label(ws_ac, r, 6, p[4])

style_cell(ws_ac, 12, 2, "Annual savings (FTE reduction):", Font(name="Aptos", size=12, bold=True, color=GREEN_ACCENT), fill_surface)
cell = ws_ac.cell(row=12, column=3, value="=(C5-C6)*C3*12")
cell.font = Font(name="Aptos", size=11, bold=True, color=GREEN_ACCENT)
cell.fill = fill_input; cell.alignment = align_center; cell.border = thin_border
cell.number_format = '#,##0'

style_cell(ws_ac, 13, 2, "Net 3-year benefit:", Font(name="Aptos", size=12, bold=True, color=GREEN_ACCENT), fill_surface)
cell = ws_ac.cell(row=13, column=3, value="=C12*3-(C9+C10*3)")
cell.font = Font(name="Aptos", size=11, bold=True, color=GREEN_ACCENT)
cell.fill = fill_input; cell.alignment = align_center; cell.border = thin_border
cell.number_format = '#,##0'

# =====================================================================
# SHEET 8: Action Plan
# =====================================================================
ws_ap = wb.create_sheet("Action Plan")
ws_ap.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_ap, [3, 8, 35, 20, 20, 20, 18])

make_header_row(ws_ap, 1, ["", "Phase", "Action Item", "Owner", "Target Date", "Status", "Notes"])

phase_data = [
    ("1", "Assess current state — complete cash app maturity score"),
    ("1", "Configure ERP auto-match rules"),
    ("1", "Enable EDI 820 remittance processing"),
    ("2", "Implement lockbox automation"),
    ("2", "Deploy deduction coding workbench"),
    ("2", "Set up unapplied cash workflow"),
    ("3", "Deploy AI-based remittance matching"),
    ("3", "Integrate customer portal auto-fetch"),
    ("3", "Go-live: straight-through processing target"),
    ("4", "Monitor KPIs and optimize thresholds"),
    ("4", "Monthly business review cadence"),
]

for r, (phase, action) in enumerate(phase_data, 2):
    style_cell(ws_ap, r, 1, "", fill_bg)
    make_label(ws_ap, r, 2, phase)
    make_label(ws_ap, r, 3, action)
    for c in range(4, 7):
        make_editable(ws_ap, r, c, "")
    make_editable(ws_ap, r, 7, "")

# =====================================================================
# SHEET 9: Assumptions
# =====================================================================
ws_as = wb.create_sheet("Assumptions & Notes")
ws_as.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_as, [3, 40, 40])

make_section_header(ws_as, 1, 2, "Model Assumptions & Key Notes", 2)
make_header_row(ws_as, 2, ["", "Assumption / Note", "Value / Detail"])

assumptions = [
    ("FTE monthly cost (fully loaded): €4,500", "Standard SSC/GBS rate W. Europe"),
    ("Benchmark source: APQC PCF v8.0 Open Standards Benchmarking 2026", "Accounts Receivable process"),
    ("Auto-cash rate baseline: 60%", "Industry median for non-automated teams"),
    ("STP definition: end-to-end without manual touch", "Includes remittance match + ERP posting"),
    ("Deduction reason codes: follow APQC PCF v8.0 taxonomy", "OpenAR standard coding"),
    ("Working days per month: 20", "For FTE capacity calculation"),
    ("Implementation cost estimate based on mid-market ERP", "SAP S/4HANA / Microsoft Dynamics"),
]

for r, (a, d) in enumerate(assumptions, 3):
    style_cell(ws_as, r, 1, "", fill_bg)
    make_label(ws_as, r, 2, a)
    make_label(ws_as, r, 3, d)

# =====================================================================
# SHEET 10: Version History
# =====================================================================
ws_vh = wb.create_sheet("Version History")
ws_vh.sheet_properties.tabColor = HEADER_BLUE
set_col_widths(ws_vh, [3, 18, 25, 20, 40])

make_header_row(ws_vh, 1, ["", "Version", "Date", "Author", "Changes"])

versions = [
    ("1.0", "2026-05-25", "ClearLedger", "Initial release — APQC PCF v8.0 aligned"),
    ("1.1", "", "", ""),
]
for r, (v, d, a, c) in enumerate(versions, 2):
    style_cell(ws_vh, r, 1, "", fill_bg)
    for i, val in enumerate([v, d, a, c], 2):
        make_editable(ws_vh, r, i, val)

# =====================================================================
# SHEET 11: Unapplied Cash (hidden)
# =====================================================================
ws_uc = wb.create_sheet("Unapplied Cash")
ws_uc.sheet_properties.tabColor = HEADER_BLUE
ws_uc.sheet_state = "hidden"
set_col_widths(ws_uc, [3, 30, 18, 18, 18, 18])

make_header_row(ws_uc, 1, ["", "Reason", "Amount (€)", "Age (days)", "Owner", "Resolution Plan"])

uc_reasons = [
    "Missing remittance advice",
    "Mismatch on invoice number",
    "Overpayment / short payment",
    "Currency mismatch",
    "Bank account not identified",
    "No matching open invoice",
]
for r, reason in enumerate(uc_reasons, 2):
    style_cell(ws_uc, r, 1, "", fill_bg)
    make_label(ws_uc, r, 2, reason)
    for c in range(3, 7):
        make_editable(ws_uc, r, c, "")

ws_uc.cell(row=9, column=2, value="Total unapplied:").font = Font(name="Aptos", size=11, bold=True, color=LABEL)
ws_uc.cell(row=9, column=2).fill = fill_surface
ws_uc.cell(row=9, column=2).border = thin_border
cell = ws_uc.cell(row=9, column=3, value="=SUM(C2:C7)")
cell.font = Font(name="Aptos", size=11, bold=True, color=EDIT_BLUE)
cell.fill = fill_input; cell.alignment = align_center; cell.border = thin_border
cell.number_format = '#,##0.00'

# =====================================================================
# SHEET 12: Mapping Table (hidden)
# =====================================================================
ws_mt = wb.create_sheet("Mapping Table")
ws_mt.sheet_properties.tabColor = HEADER_BLUE
ws_mt.sheet_state = "hidden"
set_col_widths(ws_mt, [3, 25, 25, 20])

make_header_row(ws_mt, 1, ["", "Field", "ERP Mapping", "Notes"])

mappings = [
    ("Invoice number", "BSEG-BELNR / VBRK-VBELN", "SAP field reference"),
    ("Customer ID", "BSEG-KUNNR / KNA1-KUNNR", ""),
    ("Amount", "BSEG-DMBTR", "In local currency"),
    ("Date", "BSEG-BUDAT", "Posting date"),
    ("Cash discount", "BSEG-SKFBT", ""),
    ("Payment reference", "FEBEP-AUGBL", "Clearing document"),
    ("Bank account", "BSEG-HKONT", "GL account"),
]
for r, (fld, erp, note) in enumerate(mappings, 2):
    style_cell(ws_mt, r, 1, "", fill_bg)
    make_label(ws_mt, r, 2, fld)
    make_label(ws_mt, r, 3, erp)
    make_label(ws_mt, r, 4, note)

# ── save ──
fp = r"C:\Users\nazim\Desktop\ClearLedger\templates\ClearLedger_Cash_Application_Process_Pack.xlsx"
wb.save(fp)
print(f"OK Saved: {fp}")
